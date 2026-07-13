"""
============================================================
SCCP-FR
generate_master_database.py

Auteur : Jean-Michel
Projet : SCCP-FR

Devise :
Apprendre • Construire • Partager
Rendre Star Citizen plus accessible.
============================================================
"""

# -----------------------------
# Affichage du démarrage
# -----------------------------

print("==========================================")
print("        SCCP MASTER DATABASE")
print("==========================================")
print()

print("Bonjour Jean-Michel !")
print("Bienvenue dans SCCP-FR")
print()

print("Début de la génération...")
# Importation de la bibliothèque Excel
from openpyxl import Workbook

# ====================================================
# LISTE DES COLONNES
# ====================================================

colonnes_commandes = [
    "ID",
    "Domaine",
    "Module",
    "Sous-module",
    "Nom officiel EN",
    "Nom SCCP FR",
    "Nom XML",
    "Description",
    "Raccourci clavier",
    "Périphérique",
    "Fréquence",
    "Niveau",
    "Statut",
    "Version SC",
    "Source",
    "Dernière vérification",
    "Testeur",
    "Profil",
    "Page Stream Deck",
    "Position",
    "Icône",
    "Commentaires"
]
# ====================================================
# FONCTIONS
# ====================================================


def creer_entetes(sheet, colonnes):
    """Crée automatiquement les en-têtes d'une feuille."""

    for numero_colonne, titre in enumerate(colonnes, start=1):
        sheet.cell(row=1, column=numero_colonne).value = titre

    print(f"En-têtes créés pour : {sheet.title}")

   creer_feuilles(workbook)

feuille_commandes = workbook["01_COMMANDES"]
creer_entetes(feuille_commandes, colonnes_commandes)





print()

print("Création du classeur Excel...")

# Création du classeur
workbook = Workbook()

# ==========================================
# Création des feuilles
# ==========================================

print("Création des feuilles...")

# Renommer la première feuille
def creer_feuilles(workbook):
    """Crée toutes les feuilles de la SCCP Master Database."""

    sheet = workbook.active
    sheet.title = "00_PROJET"

    feuilles = [
        "01_COMMANDES",
        "02_CATEGORIES",
        "03_PERIPHERIQUES",
        "04_TRADUCTIONS",
        "05_PROFILS",
        "06_STREAMDECK",
        "07_ICONES",
        "08_TESTS",
        "09_CHANGELOG"
    ]

    for nom in feuilles:
        workbook.create_sheet(title=nom)

    print("Toutes les feuilles ont été créées.")

print("Toutes les feuilles ont été créées.")

# Sauvegarde dans le dossier Database
workbook.save("database/SCCP_Master_Database_v0.1.xlsx")

print()
print("Base de données créée avec succès !")
print("Fin du programme.")
