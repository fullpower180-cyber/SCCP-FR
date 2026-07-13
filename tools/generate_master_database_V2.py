"""
============================================================
SCCP-FR
generate_master_database.py

Auteur : Jean-Michel
Projet : SCCP-FR

Apprendre • Construire • Partager
Rendre Star Citizen plus accessible.
============================================================
"""

# ==========================================================
# IMPORTS
# ==========================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ==========================================================
# CONSTANTES
# ==========================================================

COLONNES_COMMANDES = [
    "ID",
    "Domaine",
    "Module",
    "Sous-module",
    "Nom officiel EN",
    "Nom SCCP FR",
    "Nom XML",
    "Description",
    "Raccourci clavier",
    "Périphérique",
    "Fréquence",
    "Niveau",
    "Statut",
    "Version SC",
    "Source",
    "Dernière vérification",
    "Testeur",
    "Profil",
    "Page Stream Deck",
    "Position",
    "Icône",
    "Commentaires"
]

# ==========================================================
# FONCTIONS
# ==========================================================

def creer_feuilles(workbook):
    """Création des feuilles du classeur."""

    feuille = workbook.active
    feuille.title = "00_PROJET"

    feuilles = [
        "01_COMMANDES",
        "02_CATEGORIES",
        "03_PERIPHERIQUES",
        "04_TRADUCTIONS",
        "05_PROFILS",
        "06_STREAMDECK",
        "07_ICONES",
        "08_TESTS",
        "09_CHANGELOG"
    ]

    for nom in feuilles:
        workbook.create_sheet(title=nom)

    print("Toutes les feuilles ont été créées.")


def creer_entetes(sheet, colonnes):
    """Création des en-têtes avec mise en forme SCCP."""

    # Style SCCP
    fond_bleu = PatternFill(fill_type="solid", fgColor="1F4E78")
    police = Font(bold=True,color="FFFFFF")
    alignement = Alignment(horizontal="center",vertical="center")

    # Création des colonnes
    for numero, titre in enumerate(colonnes, start=1):

        cellule = sheet.cell(row=1, column=numero)

        cellule.value = titre
        cellule.fill = fond_bleu
        cellule.font = police
        cellule.alignment = alignement

    # Hauteur de la première ligne
    sheet.row_dimensions[1].height = 24

    print(f"En-têtes créés : {sheet.title}")

# ==========================================================
# PROGRAMME PRINCIPAL
# ==========================================================

print("==========================================")
print("        SCCP MASTER DATABASE")
print("==========================================")
print()

print("Bonjour Jean-Michel !")
print("Bienvenue dans SCCP-FR")
print()

print("Création du classeur Excel...")

workbook = Workbook()

creer_feuilles(workbook)

feuille_commandes = workbook["01_COMMANDES"]

creer_entetes(feuille_commandes, COLONNES_COMMANDES)

print(feuille_commandes["A1"].value)
print(feuille_commandes["B1"].value)

workbook.save("Database/SCCP_Master_Database_v0.1.xlsx")

print()
print("Base de données créée avec succès !")
print("Fin du programme.")